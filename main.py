import asyncio
import threading
import tempfile
import pandas as pd
import openpyxl
import os
from openpyxl.utils.dataframe import dataframe_to_rows
from fastapi import FastAPI, HTTPException
from fastapi.responses import FileResponse
from controller.scripts.headers import *
from database.queries import *

from pydantic import BaseModel
from controller.service.worker import worker

app = FastAPI()

def background_task():
    asyncio.run(createDB())
    asyncio.run(worker())

thread = threading.Thread(target=background_task)
thread.start()
'''getVT("example.com")
getHeaders("example.com")
getWhois("example.com")
getPing("example.com")
getQuality("example.com")
getSSL("example.com")
getDns("example.com")
getTextSecurity("example.com")
getTraceRoute("example.com")
getSocialTags("example.com")
getCookies("example.com")
getHttpSecurity("example.com")
getTechStack("example.com")
getArchives("example.com")
getBlockLists("example.com")
getCarbon("example.com")
getDnssec("example.com")
getDnsServer("example.com")
getDomain("example.com")
getFeatures("example.com")
getfirewall("example.com")
gethsts("example.com")
getIP("example.com")
getLinkedPages("example.com")
getMailConfig("example.com")
getRank("example.com")
getRedirects("example.com")
getRobots("example.com")
getScreenshot("example.com")
getSitemap("example.com")
getStatus("example.com")
getThreats("example.com")
getTLS("example.com")
getTxtRecords("example.com")'''
#getInternetNL()

@app.get("/")
async def root():
  #await createDB()
  a = await get_entities()
  print(a)
  print("Added!!")
  return {"message": "Message from backend"}

@app.post("/test")
def test(data):
  print(data)
  return 200



class Data(BaseModel):
    entitiesList: list

@app.post("/createEntities")
async def createEntities(data: Data):
  try:
    await create_Entity_v2(data.entitiesList)
    print("Added!!")
    return 200
  except Exception as e:
    print(e)
    raise HTTPException(status_code=500, detail="Internal server error")
  
class GroupEntitiesData(BaseModel):
    entities: list
    tag: str | None = None
    groupName: str | None = None
    createGroup: bool = True

@app.post("/createGroupEntities")
async def createGroupEntities(data: GroupEntitiesData):
  try:
    print(f"Received request to create entities: {data.entities}, {data.tag}, {data.groupName}")
    await create_Group_Entities(data.entities, data.tag, data.groupName, data.createGroup)
    print("Added!!")
    return 200
  except Exception as e:
    print(e)
    raise HTTPException(status_code=500, detail="Internal server error")
  
class GroupData(BaseModel):
    tag: str

@app.post("/createGroup")
async def createGroup(data: GroupData):
  try:
    print(f"Received request to create group: {data.tag}")
    await create_group(data.tag)
    print("Group Added!!")
    return 200
  except Exception as e:
    print(e)
    raise HTTPException(status_code=500, detail="Internal server error")
  
class RelationData(BaseModel):
    entities: list
    tag: str
    version: int
  
@app.post("/createGroupEntitiesRelation")
async def createGroupEntitiesRelation(data: RelationData):
  try:
    await create_Group_Entities_Relation(data.entities, data.tag, data.version)
    print("Added!!")
    return 200
  except Exception as e:
    print(e)
    raise HTTPException(status_code=500, detail="Internal server error")
  
@app.post("/removeGroupEntitiesRelation")
async def removeGroupEntitiesRelation(data: RelationData):
  try:
    await remove_Group_Entities_Relation(data.entities, data.tag, data.version)
    print("Added!!")
    return 200
  except Exception as e:
    print(e)
    raise HTTPException(status_code=500, detail="Internal server error")
  
  
class EntityUpdateData(BaseModel):
    url: str
    new_name: str
    new_acronym: str
    new_url: str

@app.post("/updateEntity")
async def updateEntity(data: EntityUpdateData):
  try:
    print(f"Received request to update entities: {data.new_name}, {data.new_acronym}, {data.new_url}")
    await update_entity(data.url, data.new_name, data.new_acronym, data.new_url)
    print("Updated!!")
    return 200
  except Exception as e:
    print(e)
    raise HTTPException(status_code=500, detail="Internal server error")

@app.get("/getEntities")
async def getEntities():
  result = await get_entities()
  return {result}

@app.get("/getGroups")
async def getGroups():
  result = await get_groups()
  return {result}

@app.get("/getEntityGroups")
async def getGroups(acronym):
  result = await get_entity_groups(acronym)
  return {result}

@app.get("/getGroup")
async def getGroups(tag):
  result = await get_group(tag)
  return {result}

@app.get("/getEntitiesWithout")
async def getEntitiesWithout(tag):
  result = await get_entities_without(tag)
  return {result}

@app.get("/getTests")
async def getTests():
  result = await get_tests()
  return {result}

@app.get("/getGroupVersions")
async def getGroupVersions(tag):
  result = await get_group_versions(tag)
  print(result)
  return {result}

@app.get("/getCampaigns")
async def getCampaigns():
  result = await get_campaigns()
  return {result}

@app.get("/getGroupCampaigns")
async def getGroupCampaigns(tag):
  result = await get_group_campaigns(tag)
  return {result}

@app.get("/getCampaignTests")
async def getCampaignTests(tag):
  result = await get_campaign_tests(tag)
  out = json.dumps(result, indent=4, default=str)
  return {out}

@app.get("/getCampaignResults")
async def getCampaignResults(tag):
  result = await get_campaign_results(tag)
  return {result}

@app.get("/getUnusedCampaignTests")
async def getCampaigns(tag):
  result = await get_unused_campaign_tests(tag)
  out = json.dumps(result, indent=4, default=str)
  return {out}

class CampaignData(BaseModel):
    testsId: list
    campaignTag: str
    campaignType: str
    dates: list
    groupId: str
    groupVersion: int
  
@app.post("/createCampaign")
async def createCampaign(data: CampaignData):
  try:
    print(f"Received request to create campaign: {data.testsId}, {data.groupVersion}, {data.campaignTag}, {data.campaignType}, {data.dates}, {data.groupId}")
    await create_Campaign(data.testsId, data.campaignTag, data.campaignType, data.dates, data.groupId, data.groupVersion)
    print("Campaign Created!!")
    return 200
  except Exception as e:
    print(e)
    raise HTTPException(status_code=500, detail="Internal server error")
  
class TestsData(BaseModel):
    tests: list
    tag: str

@app.post("/createNewTestsServicesVersion")
async def createNewTestsServicesVersion(data: TestsData):
  try:
    print(f"Received request to create new tests group version: {data.tests}, {data.tag}")
    await create_new_tests_services_version(data.tests, data.tag)
    return 200
  except Exception as e:
    print(e)
    raise HTTPException(status_code=500, detail="Internal server error")
  
@app.post("/removeTestsServices")
async def removeTestsServices(data: TestsData):
  try:
    print(f"Received request to remove tests: {data.tests}, {data.tag}")
    await remove_tests_services(data.tests, data.tag)
    return 200
  except Exception as e:
    print(e)
    raise HTTPException(status_code=500, detail="Internal server error")

@app.get("/getCampaignResultFile")
async def getCampaignResultFile(path):
  try:
    # Assuming the file_path is a relative path to the file
    # Adjust the file_path as needed to locate the file on your server
    with open(path, 'r') as json_file:
      # Create a temporary directory to store the Excel file
      temp_dir = tempfile.mkdtemp(dir='outputs')
      excel_file_path = os.path.join(temp_dir, 'data.xlsx')
      wb = openpyxl.Workbook()
      wb.save(excel_file_path)
      # Load JSON data
      json_data = json.load(json_file)
      for sheet_name, data in json_data.items():
        # Convert JSON data to pandas DataFrame
        df = pd.DataFrame(data)
        sheet = wb.create_sheet(title=sheet_name)
         # Write column headers with formatting
        for col_idx, column_name in enumerate(df.columns, start=1):
          header_cell = sheet.cell(row=1, column=col_idx)
          header_cell.value = column_name
          header_cell.font = openpyxl.styles.Font(bold=True)
          header_cell.alignment = openpyxl.styles.Alignment(horizontal='center')

        # Write data rows
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
          for c_idx, value in enumerate(row, start=1):
            cell = sheet.cell(row=r_idx, column=c_idx)
            cell.value = str(value)
      
      wb.remove(wb['Sheet'])
      wb.save(excel_file_path)
      # Check if the Excel file exists
      if os.path.exists(excel_file_path):
        # Return the Excel file as a FileResponse
        headers = {'Content-Disposition': 'attachment; filename="Book.xlsx"', 'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'}
        return FileResponse(excel_file_path, headers=headers, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename='data.xlsx')
      else:
        raise HTTPException(status_code=500, detail="Failed to create Excel file.")
  except Exception as e:
      print(e)
      raise HTTPException(status_code=500, detail=str(e))